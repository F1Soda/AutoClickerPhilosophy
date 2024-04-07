try:
    import selenium.common.exceptions
    from selenium.webdriver.common.by import By
    from openpyxl import load_workbook
    from selenium import webdriver
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as ec
    from enum import Enum

    import re
    import pickle
    import time
    import tkinter as tk
    import threading
    import keyboard
except:
    import pip

    pip.main(['install', '--quiet', 'selenium'])
    pip.main(['install', '--quiet', 'openpyxl'])
    pip.main(['install', '--quiet', 'enum'])
    pip.main(['install', '--quiet', 're'])
    pip.main(['install', '--quiet', 'pickle'])
    pip.main(['install', '--quiet', 'tkinter'])
    pip.main(['install', '--quiet', 'threading'])
    pip.main(['install', '--quiet', 'keyboard'])

    import selenium.common.exceptions
    from selenium.webdriver.common.by import By
    from openpyxl import load_workbook
    from selenium import webdriver
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as ec
    from enum import Enum

    import re
    import pickle
    import time
    import tkinter as tk
    import threading
    import keyboard


def parse_excel_file(path):
    book = load_workbook(filename=path)
    sheet = book['Философия']
    res_dict = {}
    i = 1
    current_theme = ''
    while i < 450:
        line = str(sheet[f'A{i}'].value)
        if line.startswith('Тема'):
            current_theme = re.search(r"(?<=:).+", line).group().strip().lower()
            res_dict[current_theme] = {}
            i += 2
            continue
        else:
            if re.fullmatch(r"\d\.\d:", line) is None:
                res_dict[current_theme][line] = sheet[f'B{i}'].value
            i += 1
    with open(r'dictionary_with_answers.txt', 'wb') as f:
        pickle.dump(res_dict, f)

    return res_dict


class BehaviourEndFillingAnswers(Enum):
    wait_until_pressed_key = 2
    send = 1
    do_nothing = 0


class AvailableBrowsers(Enum):
    firefox = webdriver.Firefox
    chrome = webdriver.Chrome
    edge = webdriver.Edge


class EventDelegate:
    def __init__(self):
        self.__event_followers = []

    def __iadd__(self, ehandler):
        self.__event_followers.append(ehandler)
        return self

    def __isub__(self, ehandler):
        self.__event_followers.remove(ehandler)
        return self

    def __call__(self, *args, **kwargs):
        for follower in self.__event_followers:
            follower(*args, **kwargs)


class WebAPI:
    def __init__(self, waiting_time):
        self.driver = None
        self.waiting_time = waiting_time
        self.found_page_with_test_event = EventDelegate()
        self.initialized_event = EventDelegate()
        self.end_filling_test_event = EventDelegate()
        self.can_not_filling_test_event = EventDelegate()
        self.on_find_page_with_questions = EventDelegate()
        self.log_event = EventDelegate()
        self.event_stop_thread = threading.Event()

    def initialize(self, password, login, auto_filling_mode, link, browser):
        self.driver = browser()
        self.driver.get(link)
        self.driver.implicitly_wait(self.waiting_time)
        if auto_filling_mode:
            self.driver.find_element(By.XPATH,
                                     value='//*[@id="root"]/div/main/div/div[1]/div[2]/form/label[1]/input').send_keys(
                login)
            self.driver.find_element(By.XPATH,
                                     value='//*[@id="root"]/div/main/div/div[1]/div[2]/form/label[2]/input').send_keys(
                password)
            self.driver.find_element(By.XPATH, value='//*[@id="root"]/div/main/div/div[1]/div[2]/form/button').click()
            self.initialized_event()
        else:
            WebDriverWait(self.driver, 60).until(
                ec.presence_of_element_located((By.XPATH, '//div[@class="position-relative"]')))
            self.initialized_event()

    def start_answering(self, answers, duration_sleep):
        self.driver.switch_to.default_content()
        title = self.driver.find_element(By.XPATH,
                                         value="//a[@class='text-primary-500'][contains(text(), 'Тема')]").text
        test_module = self.driver.find_element(By.XPATH, "//h1[@class='mb-0 h3']").text
        if test_module == 'Тест по материалам лекции 10':
            frame = self.driver.find_element(by=By.XPATH, value="//iframe[@id='unit-iframe']")
            self.driver.switch_to.frame(frame)
            self.driver.find_element(By.XPATH, '//label[contains(text(),"Сэмюэля Александера")]').click()
            self.driver.find_element(By.XPATH, '//label[contains(text(),"Моргана Конви Ллойда")]').click()
            self.driver.find_element(By.XPATH, '//label[contains(text(),"принцип всеобщей связи")]').click()
            self.driver.find_element(By.XPATH, '//label[contains(text(),"принцип становления")]').click()
            self.driver.find_element(By.XPATH, '//label[contains(text(),"принцип историзма")]').click()
            self.driver.switch_to.default_content()

        self.driver.implicitly_wait(1)
        is_independent_work = len(
            self.driver.find_elements(By.XPATH, "//h1[@class='mb-0 h3'][contains(text(),'Самостоятельная')]")) == 1
        theme = ' '.join(title.split()[2:]).lower()
        theme_answers = answers[theme]
        frame = self.driver.find_element(by=By.XPATH, value="//iframe[@id='unit-iframe']")
        self.driver.switch_to.frame(frame)

        questions = self.driver.find_elements(By.XPATH,
                                              value='//legend[@class="response-fieldset-legend field-group-hd"]')

        for question in questions:
            question = question.text
            if self.event_stop_thread.is_set():
                break
            answer = theme_answers[question]
            for answer in answer.split('\n'):
                self.driver.find_element(By.XPATH,
                                         value=self.__generate_xpath_for_point_question(question, answer)).click()
            time.sleep(duration_sleep)

        input_fields = self.driver.find_elements(By.XPATH, value='//input[@type="text"]')

        if len(input_fields) != 0:
            count_processed_input_filed = 0
            for q_a in self.get_last_answers(theme_answers, len(input_fields)):
                if q_a[0] == 'Каково назначение ученого, по Фихте?':
                    input = self.driver.find_element(By.XPATH, '//input[@type="text"]')
                    input.clear()
                    input.send_keys('абстрактным')
                    continue

                if self.event_stop_thread.is_set():
                    break
                if q_a[0] not in questions:
                    input_fields[count_processed_input_filed].clear()
                    input_fields[count_processed_input_filed].send_keys(q_a[1])
                    count_processed_input_filed += 1

        p_blocks = self.driver.find_elements(By.XPATH, value='//div[@class="problem"]//p')

        if len(p_blocks) != 0 and not is_independent_work:
            for i in range(len(p_blocks)):
                if p_blocks[i].text in theme_answers:
                    answer = theme_answers[p_blocks[i].text]
                    for ans in answer.split('\n'):
                        if self.event_stop_thread.is_set():
                            break
                        self.driver.find_element(By.XPATH,
                                                 self.__generate_xpath_for_point_question_p_tag(ans,
                                                                                                i + 1)).click()

        if len(questions) == 0 and is_independent_work:
            try:
                questions = self.driver.find_elements(By.XPATH, value='//div[@class="wrapper-problem-response"]')
                if len(questions) != 0:
                    q_as = self.get_last_answers(theme_answers, len(questions))
                    for i in range(len(q_as)):
                        if self.event_stop_thread.is_set():
                            break
                        self.driver.find_element(By.XPATH,
                                                 self.__generate_xpath_for_point_question_wrapper_problem(
                                                     q_as[i][1])).click()
            except selenium.common.exceptions.NoSuchElementException:
                pass

        self.driver.implicitly_wait(self.waiting_time)
        self.end_filling_test_event()

    def get_last_answers(self, theme_answers, count):
        self.driver.implicitly_wait(self.waiting_time)
        i = 0
        res_answers = []
        answers = theme_answers.items()
        for answer in answers:
            if i >= len(answers) - count - 1:
                if answer[1] is None:
                    continue
                res_answers.append(answer)
            i += 1
        return res_answers

    def __generate_xpath_for_point_question(self, question, answer):
        # //legend[@class='response-fieldset-legend field-group-hd'][contains(text(), '1.')][contains(text(), 'Комплекс')][contains(text(), 'вопросов')]/..//label[contains(text(), 'неклассической')]
        res = "//legend[@class='response-fieldset-legend field-group-hd']"
        key_word = question.split()
        count_iteration = 10 if len(key_word) > 10 else len(key_word)
        for word in key_word[:count_iteration]:
            res += f"[contains(text(), '{word}')]"
        res += "/..//label"
        key_word = answer.split()
        count_iteration = 5 if len(key_word) > 5 else len(key_word)
        index = 2
        count_inner_iterations = 0
        for word in key_word[:count_iteration]:
            res += f"[starts-with(substring(text(), {index}, {100}), '{word}')"
            for i in range(1, count_inner_iterations + 1):
                res += f"or starts-with(substring(text(), {index + i}, {100}), '{word}')"
            res += ']'
            count_inner_iterations += 1
            index += len(word) + 1
        return res

    def __generate_xpath_for_point_question_p_tag(self, answer, index):
        res = f"//div[@class='problem']//p[{index}]//following::div[1]"
        res += "//label"
        key_word = answer.split()
        count_iteration = 7 if len(key_word) > 7 else len(key_word)
        index = 2
        count_inner_iterations = 0
        for word in key_word[:count_iteration]:
            res += f"[starts-with(substring(text(), {index}, {100}), '{word}')"
            for i in range(1, count_inner_iterations + 1):
                res += f"or starts-with(substring(text(), {index + i}, {100}), '{word}')"
            res += ']'
            count_inner_iterations += 1
            index += len(word) + 1
        return res

    def __generate_xpath_for_point_question_wrapper_problem(self, answer):
        res = "//div[@class='wrapper-problem-response']//label"
        key_word = answer.split()
        count_iteration = 5 if len(key_word) > 5 else len(key_word)
        for word in key_word[:count_iteration]:
            res += f"[contains(text(), '{word}')]"
        return res

    def __generate_xpath_for_input_field_question(self, question):
        res = '//p'
        key_word = question.split()
        count_iteration = 5 if len(key_word) > 5 else len(key_word)
        for word in key_word[:count_iteration]:
            res += f"[contains(text(), '{word}')]"
        res += '/..//input'
        return res

    def try_send_answer(self):
        self.driver.implicitly_wait(0.1)
        counter = 0
        buttons_next = self.driver.find_elements(By.XPATH, '//button[@data-submitting="Отправка"]')
        self.driver.implicitly_wait(self.waiting_time)
        for button in buttons_next:
            if button.get_attribute('data-should-enable-submit-button') == 'False':
                continue
            else:
                button.click()
                counter += 1
        return counter == len(buttons_next)

    def try_find_block_questions(self, make_page_green):
        t_end = time.time() + 30
        self.driver.implicitly_wait(0.5)
        self.driver.switch_to.default_content()
        while t_end > time.time():
            if self.event_stop_thread.is_set():
                self.driver.switch_to.default_content()
                self.log_event('Остановка автозаполнения')
                return None
            try:
                frame = self.driver.find_element(by=By.XPATH, value="//iframe[@id='unit-iframe']")
                self.driver.switch_to.frame(frame)
                self.driver.find_element(By.XPATH, value='//div[@class="problem"]')
            except selenium.common.exceptions.NoSuchElementException:
                if make_page_green:
                    self.make_page_green()
                self.load_next_page()
            else:
                self.driver.implicitly_wait(self.waiting_time)
                self.found_page_with_test_event()
                return True
        self.driver.switch_to.default_content()
        self.log_event('Невозможно найти страницу с тестом')
        return False

    def make_page_green(self):
        self.driver.implicitly_wait(5)

        play_video_buttons = self.driver.find_elements(By.XPATH,
                                                       "//button[@class='plyr__control plyr__control--overlaid']")
        if len(play_video_buttons) != 0:
            self.driver.find_element(By.XPATH, "//button[@class='plyr__control'][@data-plyr='mute']").click()
            play_video_buttons[0].click()
            time.sleep(4)
            return None
        self.driver.implicitly_wait(0.1)

        practice_task = self.driver.find_elements(By.XPATH, '//h1[@class="mb-0 h3"][contains(text(), "Практическое")]')
        if len(practice_task) != 0:
            return None



        links = self.driver.find_elements(By.XPATH,
                                          "//a[contains(text(), 'Конспект') or contains(text(), 'Презентация')]")
        if len(links) != 0:
            self.driver.execute_script("window.open('');")
            link = links[0].get_attribute('href')
            self.driver.switch_to.window(self.driver.window_handles[1])
            self.driver.get(link)
            time.sleep(5)
            self.driver.close()
            self.driver.switch_to.window(self.driver.window_handles[0])
            return None

        return None

    def load_next_page(self):
        self.driver.switch_to.default_content()
        self.driver.find_element(By.XPATH, value="//button[@class='next-btn btn btn-link']").click()

    def quit(self):
        if self.driver is not None:
            self.driver.quit()


class App(tk.Tk):
    def __init__(self, web_api, link=''):
        super().__init__()
        self.stop_loop_waiting_answer = False
        self.show_hide_password = None
        self.joke1 = None
        self.login_field = None
        self.start_auto_filling_button = None
        self.stop_auto_filling_button = None
        self.password_field = None
        self.duration_between_answering_field = None
        self.link_field = None
        self.browser_drop_down_menu = None
        self.user_browser_StringVar = None
        self.behaviour_drop_down_menu = None
        self.current_behaviour_tk_StringVar = None
        self.console_text_box = None
        self.answers = None
        self.password = ''
        self.login = ''
        self.web_api = web_api
        self.th = None
        self.auto_filling_flag = False
        self.link = link
        self.make_page_green = False
        self.joke1 = None
        self.joke2 = None
        self.joke3 = None
        self.duration_between_answering = '0'
        self.behaviour_after_end_filling_answers = BehaviourEndFillingAnswers.do_nothing
        self.user_browser = AvailableBrowsers.edge
        self.web_api.initialized_event += self.set_button_start_auto_filling_enable
        self.web_api.end_filling_test_event += self.on_end_filling_answers
        self.web_api.can_not_filling_test_event += self.on_can_not_fill_answers
        self.web_api.found_page_with_test_event += self.start_auto_filling
        self.web_api.log_event += self.log
        self.initialize_gui()

    def load_user_data(self):
        try:
            with open(r'user_data.txt', 'rb') as f:
                data = pickle.load(f)
                self.password = data['password']
                self.login = data['login']
                self.link = data['link']
                self.auto_filling_flag = data['auto_filling_flag']
                self.duration_between_answering = data['duration_between_answering']
                self.behaviour_after_end_filling_answers = data['behaviour_after_end_filling_answers']
                self.user_browser = data['user_browser']
                self.make_page_green = data['make_page_green']
                self.show_hide_password = data['show_hide_password']
                self.joke1 = data['joke1']
                self.joke2 = data['joke2']
                self.joke3 = data['joke3']
            self.answers = self.try_get_dict_answers()
        except FileNotFoundError:
            self.log("Файл с сохранёнными данными не найден")
        except EOFError:
            self.log("Файл с пользовательскими даннами не в нужном формате")

    def save_user_data(self):
        dict_to_save = {
            'password': self.password_field.get(),
            'login': self.login_field.get(),
            'link': self.link_field.get(),
            'auto_filling_flag': self.auto_filling_flag,
            'duration_between_answering': self.duration_between_answering_field.get(),
            'behaviour_after_end_filling_answers': BehaviourEndFillingAnswers[
                self.current_behaviour_tk_StringVar.get().replace(' ', '_')],
            'user_browser': AvailableBrowsers[self.user_browser_StringVar.get()],
            'make_page_green': self.make_page_green,
            'show_hide_password': self.show_hide_password_var.get(),
            'joke1': False if self.joke1 is None else self.joke1,
            'joke2': (False, False) if self.joke2 is None else self.joke2,
            'joke3': False if self.joke3 is None else self.joke3,
        }
        with open(r'user_data.txt', 'wb') as f:
            pickle.dump(dict_to_save, f)

    def log(self, message):
        self.console_text_box['state'] = tk.NORMAL
        self.console_text_box.insert(tk.END, str(message) + '\n')
        self.console_text_box['state'] = tk.DISABLED

    def log_letter_by_letter(self, message, checkbox):
        self.console_text_box['state'] = tk.NORMAL
        for letter in str(message):
            self.console_text_box.insert(tk.END, letter)
            time.sleep(0.025)
        self.console_text_box.insert(tk.END, '\n')
        if checkbox is not None:
            checkbox.set(False)
        self.console_text_box['state'] = tk.DISABLED

    def initialize_gui(self):
        self.console_text_box = tk.Text(self, height=10, width=60, state=tk.DISABLED)
        self.console_text_box.grid(row=8, column=0, columnspan=3)
        self.log("Логгер")
        self.load_user_data()

        tk.Label(self, text="Ссылка на страничку").grid(row=0)
        tk.Label(self, text="Логин").grid(row=1)
        tk.Label(self, text="Пароль").grid(row=2)
        tk.Label(self, text="Задержка ввода в секундах").grid(row=3)
        tk.Label(self, text="Поведение после заполнения").grid(row=4)
        tk.Label(self, text="Запускаемый браузер").grid(row=7)

        self.current_behaviour_tk_StringVar = tk.StringVar()
        self.current_behaviour_tk_StringVar.set(self.behaviour_after_end_filling_answers.name.replace('_', ' '))
        self.behaviour_drop_down_menu = tk.OptionMenu(self, self.current_behaviour_tk_StringVar,
                                                      *[beh.name.replace('_', ' ') for beh in
                                                        BehaviourEndFillingAnswers])
        self.behaviour_drop_down_menu.grid(row=4, column=1)

        self.user_browser_StringVar = tk.StringVar()
        self.user_browser_StringVar.set(self.user_browser.name)
        self.browser_drop_down_menu = tk.OptionMenu(self, self.user_browser_StringVar,
                                                    *[browser.name for browser in
                                                      AvailableBrowsers])
        self.browser_drop_down_menu.grid(row=7, column=1)

        self.link_field = tk.Entry(self)
        self.link_field.insert(tk.END, self.link)
        self.login_field = tk.Entry(self)
        self.login_field.insert(tk.END, self.login)
        self.password_field = tk.Entry(self)
        self.password_field.insert(tk.END, self.password)
        self.duration_between_answering_field = tk.Entry(self)
        self.duration_between_answering_field.insert(tk.END, self.duration_between_answering)

        self.link_field.grid(row=0, column=1)
        self.login_field.grid(row=1, column=1)
        self.password_field.grid(row=2, column=1)
        if self.show_hide_password:
            self.password_field.configure(show="*")
        self.duration_between_answering_field.grid(row=3, column=1)

        tk.Button(self,
                  text='Выход',
                  command=self.close_application).grid(row=18,
                                                       column=0,
                                                       sticky=tk.W,
                                                       pady=4)

        auto_filling_var = tk.IntVar()
        auto_filling_var.set(self.auto_filling_flag)
        (tk.Checkbutton(self,
                        text='Авторегистрация',
                        variable=auto_filling_var,
                        onvalue=True,
                        offvalue=False,
                        command=(lambda:
                                 self.checkbutton_auto_filling_password_changed(auto_filling_var))).grid(row=5,
                                                                                                         column=1,
                                                                                                         sticky=tk.W,
                                                                                                         pady=4))
        self.show_hide_password_var = tk.IntVar(value=self.show_hide_password)
        tk.Checkbutton(self,
                       text='Скрыть/Показать',
                       variable=self.show_hide_password_var,
                       onvalue=True,
                       offvalue=False,
                       command=(lambda: self.checkbutton_show_hide_password_changed(self.show_hide_password_var))).grid(
            row=2,
            column=2,
            sticky=tk.W,
            pady=4)
        self.set_green_page_var = tk.IntVar(value=self.make_page_green)
        tk.Checkbutton(self,
                       text='Делать странички зелёными?\n(для эстетов)',
                       variable=self.set_green_page_var,
                       onvalue=True,
                       offvalue=False,
                       command=(lambda: self.checkbutton_set_green_page_changed(self.set_green_page_var))).grid(row=6,
                                                                                                                column=1,
                                                                                                                sticky=tk.W,
                                                                                                                pady=4)

        tk.Button(self, text='Запустить браузер', command=self.run_browser).grid(row=15,
                                                                                 column=1,
                                                                                 pady=4)

        self.start_auto_filling_button = tk.Button(self, text='Начать заполнение', command=self.try_find_element,
                                                   state=tk.DISABLED)
        self.start_auto_filling_button.grid(row=16,
                                            column=1,
                                            pady=4)

        self.stop_auto_filling_button = tk.Button(self, text='Прекратить заполнение', command=self.stop_auto_filling,
                                                  state=tk.DISABLED)
        self.stop_auto_filling_button.grid(row=17,
                                           column=1,
                                           pady=4)

        self.mainloop()

    def try_find_element(self):
        self.start_auto_filling_button['state'] = tk.DISABLED
        self.stop_auto_filling_button['state'] = tk.NORMAL
        self.web_api.event_stop_thread.clear()
        self.th = threading.Thread(target=self.web_api.try_find_block_questions, args=(self.set_green_page_var.get(),))
        self.th.start()

    def start_auto_filling(self):
        self.th = threading.Thread(target=self.web_api.start_answering,
                                   args=(self.answers, (int(self.duration_between_answering_field.get())),))
        self.th.start()

    def stop_auto_filling(self):
        self.stop_auto_filling_button['state'] = tk.DISABLED
        self.start_auto_filling_button['state'] = tk.NORMAL
        if self.behaviour_drop_down_menu == BehaviourEndFillingAnswers.wait_until_pressed_key:
            self.stop_loop_waiting_answer = True
        self.web_api.event_stop_thread.set()
        self.web_api.driver.switch_to.default_content()

    def checkbutton_auto_filling_password_changed(self, value):
        self.auto_filling_flag = value.get()

    def checkbutton_show_hide_password_changed(self, value):
        if value.get():
            if not self.joke1:
                threading.Thread(target=self.log_letter_by_letter,
                                 args=(
                                     'Чего боишься? Всё равно у тебя один и тот же пароль на нескольких сайтах.',
                                     None)).start()
                self.joke1 = True
            self.password_field.configure(show="*")
        else:
            self.password_field.configure(show="")

    def checkbutton_set_green_page_changed(self, value):
        if value.get():
            self.make_page_green = True
            if not self.joke2[0]:
                threading.Thread(target=self.log_letter_by_letter,
                                 args=(
                                     'Чёрт побери, тебе настолько важно видеть зелёные галочки в бесполезном курсе? Я лучше выключу её.',
                                     self.set_green_page_var)).start()
                self.make_page_green = False
                # value.set(False)
                self.joke2 = (True, False)
            elif not self.joke2[1]:
                threading.Thread(target=self.log_letter_by_letter,
                                 args=(
                                     'Сраный эстет.', None)).start()
                self.joke2 = (True, True)
        else:
            self.make_page_green = False

    def set_button_start_auto_filling_enable(self):
        self.start_auto_filling_button['state'] = tk.NORMAL

    def on_end_filling_answers(self):
        self.behaviour_after_end_filling_answers = BehaviourEndFillingAnswers[
            self.current_behaviour_tk_StringVar.get().replace(' ', '_')]
        if self.behaviour_after_end_filling_answers is BehaviourEndFillingAnswers.do_nothing:
            self.stop_auto_filling_button['state'] = tk.DISABLED
            self.set_button_start_auto_filling_enable()
        elif self.behaviour_after_end_filling_answers is BehaviourEndFillingAnswers.wait_until_pressed_key:
            self.send_answers()
            self.log("Ожидание нажатия Enter")
            threading.Thread(target=self.wait_until_press_enter).start()
        else:
            if self.send_answers():
                self.try_find_element()

    def on_can_not_fill_answers(self, answers):
        self.log('Невозможно заполнить поля ответов. Ответы для всего модуля:')
        for answer in answers:
            self.log(answer)
        self.stop_auto_filling_button['state'] = tk.DISABLED
        self.start_auto_filling_button['state'] = tk.NORMAL

    def wait_until_press_enter(self):
        timer = time.time()
        while True:  # making a loop
            try:  # used try so that if user pressed other than the given key error will not be shown
                if keyboard.is_pressed('enter'):  # if key 'q' is pressed
                    self.log("Нажато!\n")
                    self.web_api.load_next_page()
                    self.try_find_element()
                    break  # finishing the loop
                if self.stop_loop_waiting_answer:
                    self.stop_loop_waiting_answer = False
                    break
            except:
                continue


    def send_answers(self):
        if self.web_api.try_send_answer():
            if self.behaviour_after_end_filling_answers == BehaviourEndFillingAnswers.send:
                self.web_api.load_next_page()
            return True
        else:
            self.log("Невозможно отправить ответы")
            if self.behaviour_after_end_filling_answers == BehaviourEndFillingAnswers.wait_until_pressed_key:
                self.log("Пропускаем и идём дальше?")
            else:
                self.stop_auto_filling_button['state'] = tk.DISABLED
                self.start_auto_filling_button['state'] = tk.NORMAL
            return False

    def run_browser(self):
        if self.th is not None:
            self.web_api.quit()
            self.th.join()
        self.th = threading.Thread(target=self.web_api.initialize, args=(
            (self.password_field.get()), (self.login_field.get()), self.auto_filling_flag,
            (self.link_field.get()), AvailableBrowsers[self.user_browser_StringVar.get()].value,))
        self.th.start()

    def close_application(self):
        #self.joke1 = self.joke2 = self.joke3 = None
        self.save_user_data()
        self.web_api.quit()
        self.quit()

    def try_get_dict_answers(self):
        # with open(r'dictionary_with_answers.txt', 'wb') as f:
        #     pickle.dump(parse_excel_file(r"E:\Downloads\Telegram Desktop\Filosofia.xlsx"), f)
        try:
            with open(r'dictionary_with_answers.txt', 'rb') as f:
                return pickle.load(f)
        except FileNotFoundError:
            print('Отсутсвует файл с ответами')
        return None


if __name__ == '__main__':
    web_driver = WebAPI(15)
    app = App(web_driver)
